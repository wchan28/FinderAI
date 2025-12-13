from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Union

from openpyxl import load_workbook


def extract_text_from_xlsx(file_path: Union[str, Path]) -> List[Dict]:
    """
    Extract text from an Excel spreadsheet.

    Returns a list of dicts, one per sheet:
    [
        {"page_number": 1, "text": "sheet 1 content..."},
        {"page_number": 2, "text": "sheet 2 content..."},
    ]
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True, read_only=True)

    sheets_content = []

    for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
        sheet = wb[sheet_name]
        sheet_text_parts = [f"Sheet: {sheet_name}"]

        for row in sheet.iter_rows():
            row_values = []
            for cell in row:
                if cell.value is not None:
                    cell_str = str(cell.value).strip()
                    if cell_str:
                        row_values.append(cell_str)
            if row_values:
                sheet_text_parts.append(" | ".join(row_values))

        sheet_text = "\n".join(sheet_text_parts)

        if len(sheet_text_parts) > 1:
            sheets_content.append({
                "page_number": sheet_index,
                "text": sheet_text
            })

    wb.close()
    return sheets_content


def extract_full_text_from_xlsx(file_path: Union[str, Path]) -> str:
    """
    Extract all text from an Excel file as a single string.
    Includes sheet markers for context.
    """
    sheets = extract_text_from_xlsx(file_path)

    full_text_parts = []
    for sheet in sheets:
        full_text_parts.append(sheet["text"])

    return "\n\n".join(full_text_parts)
